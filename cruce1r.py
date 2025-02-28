import pandas as pd
import openpyxl
from pathlib import Path
from collections import defaultdict
import re

def clean_id(id_value):
    """Limpia y normaliza un ID para asegurar que sea un string de 16 dígitos."""
    if pd.isna(id_value):
        return None
        
    # Convierte a string y elimina espacios y caracteres no deseados
    id_str = str(id_value).strip().replace('\xa0', '')
    
    # Elimina ".0" al final si existe (común en números leídos como float)
    if id_str.endswith('.0'):
        id_str = id_str[:-2]
    
    # Maneja notación científica
    if 'e' in id_str.lower():
        try:
            return str(int(float(id_str)))
        except:
            return id_str
    
    # Elimina caracteres no numéricos
    id_str = re.sub(r'\D', '', id_str)
    
    # Verifica si tiene exactamente 16 dígitos
    if len(id_str) == 16:
        return id_str
    else:
        return None  # No es un ID válido si no tiene 16 dígitos

def normalize_value(val):
    """Convierte un valor a cadena de forma consistente.
    Si es un número, se convierte a entero si es exacto o a float con dos decimales."""
    if val is None:
        return ""
    try:
        if isinstance(val, float):
            if val.is_integer():
                return str(int(val))
            else:
                return f"{val:.2f}"
        if isinstance(val, int):
            return str(val)
        return str(val).strip()
    except Exception:
        return str(val).strip()

def relevant_columns_signature(data_list):
    """
    Genera una 'firma' *solo* con los valores de las columnas J(10), N(14), O(15), R(18) y U(21).
    En el Excel final:
        - Columna 4 en Excel => data_list[0]
        - Columna 5 en Excel => data_list[1]
        ...
        - Columna 10 (J) => data_list[6]
        - Columna 14 (N) => data_list[10]
        - Columna 15 (O) => data_list[11]
        - Columna 18 (R) => data_list[14]
        - Columna 21 (U) => data_list[17]
    """
    # Asegurarnos de no salirnos de rango
    indices = [6, 10, 11, 14, 17]  # J, N, O, R, U
    values = []
    for idx in indices:
        if idx < len(data_list):
            values.append(normalize_value(data_list[idx]))
        else:
            # Si la fila no tiene tantas columnas, tomamos ""
            values.append("")
    return tuple(values)

def process_excel():
    base_path = Path.cwd()
    print(f"Directorio actual: {base_path}")
    
    # Procesar TODOS los Excel de REPORTE-ML (cualquier nombre, extensión .xls o .xlsx)
    reporte_files = list(base_path.joinpath('REPORTE-ML').glob('*.xls*'))
    if not reporte_files:
        print("No se encontraron archivos Excel en REPORTE-ML")
        return
    print("Archivos encontrados en REPORTE-ML:")
    for f in reporte_files:
        print(f"- {f.name}")
    
    concentrado_path = base_path.joinpath('RESULTADO-FINAL', 'CONCENTRADO-MERCADOLIBRE.xlsx')
    if not concentrado_path.exists():
        print("No se encontró el archivo CONCENTRADO-MERCADOLIBRE.xlsx")
        return
    
    print("\nProcesando archivos de reporte...")
    all_report_data = []
    invalid_files = []
    
    for reporte_path in reporte_files:
        try:
            df = pd.read_excel(reporte_path)
            
            # Verificar que la primera columna exista y contenga IDs válidos de 16 dígitos
            if df.shape[1] == 0:
                print(f"Advertencia: {reporte_path.name} no tiene columnas")
                invalid_files.append(reporte_path.name)
                continue
                
            # Verificar si hay al menos un ID válido de 16 dígitos en la columna A
            valid_ids = df.iloc[:, 0].apply(lambda x: clean_id(x) is not None).sum()
            if valid_ids == 0:
                print(f"Advertencia: {reporte_path.name} no contiene IDs válidos de 16 dígitos en la columna A")
                invalid_files.append(reporte_path.name)
                continue
                
            all_report_data.append(df)
            print(f"Leído {reporte_path.name}: {len(df)} registros, {len(df.columns)} columnas, {valid_ids} IDs válidos")
        except Exception as e:
            print(f"Error leyendo {reporte_path.name}: {e}")
            invalid_files.append(reporte_path.name)
    
    if not all_report_data:
        print("No hay archivos de reporte válidos para procesar")
        return
        
    reporte = pd.concat(all_report_data, ignore_index=True)
    print(f"\nTotal registros combinados: {len(reporte)}")
    
    # Construir diccionario de datos del reporte:
    # La primera columna es el ID y se toma toda la información restante.
    reporte_dict = defaultdict(list)
    ml_ids_not_found = set()  # Para almacenar IDs que no se encuentran en el concentrado
    
    for _, row in reporte.iterrows():
        id_value = clean_id(row.iloc[0])
        if id_value:
            reporte_dict[id_value].append({
                'data': row.iloc[1:].tolist(),
                'used': False
            })
    print(f"IDs únicos en reportes: {len(reporte_dict)}")
    
    wb_concentrado = openpyxl.load_workbook(concentrado_path)
    ws_concentrado = wb_concentrado.active

    # Recorrer las filas existentes del concentrado (a partir de la fila 2) y extraer datos desde la columna A
    existing_rows = defaultdict(list)
    max_consecutive = 0  # Para llevar el control del mayor consecutivo encontrado
    
    for row in range(2, ws_concentrado.max_row + 1):
        id_value = clean_id(ws_concentrado.cell(row=row, column=1).value)
        
        # Obtener el consecutivo (columna B)
        consecutivo = ws_concentrado.cell(row=row, column=2).value
        if isinstance(consecutivo, (int, float)) and consecutivo > max_consecutive:
            max_consecutive = consecutivo
            
        # Obtener observaciones (columna C)
        observaciones = ws_concentrado.cell(row=row, column=3).value
        
        if id_value:
            # row_data = columnas desde la 4 hasta el final
            row_data = [ws_concentrado.cell(row=row, column=col).value 
                        for col in range(4, ws_concentrado.max_column + 1)]
            existing_rows[id_value].append((row, consecutivo, observaciones, row_data))
    
    print(f"IDs en concentrado: {len(existing_rows)}")
    print(f"Mayor consecutivo encontrado: {max_consecutive}")
    
    # Verificar qué IDs de reporte_dict no existen en existing_rows
    for id_value in reporte_dict.keys():
        if id_value not in existing_rows:
            ml_ids_not_found.add(id_value)
    
    changes_made = False
    updated_rows = 0
    new_rows_count = 0
    duplicates_omitted_global = 0  # Contador de filas duplicadas omitidas
    
    # Primero actualizamos filas existentes y agregamos nuevas filas para IDs existentes
    print("\nActualizando datos y agregando filas para IDs existentes (comparando solo columnas J, N, O, R, U)...")
    for id_value, rows_info in existing_rows.items():
        if id_value not in reporte_dict:
            continue
            
        # Verificar todas las entradas en reporte_dict para este ID
        entries_to_process = reporte_dict[id_value].copy()
        
        # Si hay al menos una fila existente y entradas en el reporte,
        # actualizamos la primera fila existente con la primera entrada del reporte
        if rows_info and entries_to_process:
            first_row_num, consecutivo, observaciones, _ = rows_info[0]
            first_entry = entries_to_process[0]
            
            # Actualizar la primera fila existente con los nuevos datos
            # Modificación: iniciar la actualización desde la columna I (índice 9) en lugar de la columna D (índice 4)
            for col_idx, value in enumerate(first_entry['data'], start=9):
                cell = ws_concentrado.cell(row=first_row_num, column=col_idx, value=value)
                if isinstance(value, (int, float)):
                    if isinstance(value, float) and not value.is_integer():
                        cell.number_format = "0.00"
                    else:
                        cell.number_format = "0"
            
            # Marcar como usada la primera entrada
            reporte_dict[id_value][0]['used'] = True
            updated_rows += 1
            changes_made = True
            print(f"Actualizada fila {first_row_num} para ID {id_value}")
            
            # Procesar las entradas restantes (añadir como nuevas filas) solo si
            # en las columnas J, N, O, R, U no coinciden con alguna fila existente
            if len(entries_to_process) > 1:
                new_rows = []
                duplicates_omitted = 0

                # Construir un set de "firmas" basado en las 5 columnas relevantes
                # Para los registros existentes, la data insertada comienza a partir de la columna I
                # Dado que en el concentrado la información original se extrajo desde la columna D,
                # para validación usamos el segmento correspondiente; en este ejemplo, se asume que
                # la data a partir de la columna I se encuentra en row_data[5:] (ajustar si es necesario).
                existing_signatures = set()
                for (_, _, _, row_data) in rows_info:
                    # Suponiendo que row_data[0] corresponde a la columna D, dejamos pasar las primeras 5 columnas
                    candidate_existing = row_data[5:]
                    sig = relevant_columns_signature(candidate_existing)
                    existing_signatures.add(sig)
                
                for entry in entries_to_process[1:]:
                    if not entry['used']:
                        # Crear nueva fila con mismo ID, consecutivo y observaciones
                        # Modificación: iniciar la inserción de datos a partir de la columna I (índice 9)
                        new_row = [None] * max(ws_concentrado.max_column, 8 + len(entry['data']))
                        new_row[0] = id_value  # ID
                        new_row[1] = consecutivo  # Mismo consecutivo
                        new_row[2] = observaciones  # Mismas observaciones
                        
                        for idx, value in enumerate(entry['data']):
                            new_row[idx + 8] = value
                        
                        # De las columnas I en adelante, se extrae la data para comparar la firma
                        candidate_data = new_row[8:]
                        sig_new = relevant_columns_signature(candidate_data)
                        
                        if sig_new in existing_signatures:
                            duplicates_omitted += 1
                            print(f"Fila duplicada omitida para ID {id_value}")
                            entry['used'] = True
                        else:
                            new_rows.append(new_row)
                            existing_signatures.add(sig_new)
                            entry['used'] = True
                
                if new_rows:
                    # Insertar justo después de la última fila con el mismo ID
                    last_row = max(row for (row, _, _, _) in rows_info)
                    insert_at = last_row + 1
                    
                    ws_concentrado.insert_rows(insert_at, amount=len(new_rows))
                    for i, row_data in enumerate(new_rows):
                        target_row = insert_at + i
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = ws_concentrado.cell(row=target_row, column=col_idx, value=value)
                            # Modificación: aplicar formato desde la columna I en adelante (col_idx >= 9)
                            if col_idx >= 9 and isinstance(value, (int, float)):
                                if isinstance(value, float) and not value.is_integer():
                                    cell.number_format = "0.00"
                                else:
                                    cell.number_format = "0"
                    
                    new_rows_count += len(new_rows)
                    print(f"Para ID {id_value}, se insertaron {len(new_rows)} fila(s) a partir de la fila {insert_at}")
                    changes_made = True
                
                duplicates_omitted_global += duplicates_omitted
    
    # Generar informe de IDs no encontrados
    if ml_ids_not_found:
        print("\n=== REPORTE DE IDS NO ENCONTRADOS EN CONCENTRADO ===")
        print(f"Se encontraron {len(ml_ids_not_found)} IDs en archivos de reporte que NO existen en el concentrado.")
        print("Estos IDs no fueron procesados:")
        for id_value in sorted(ml_ids_not_found):
            print(f"- {id_value} (aparece {len(reporte_dict[id_value])} veces en los reportes)")
        print("=======================================================")

    if changes_made:
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        backup_path = concentrado_path.parent / f"CONCENTRADO-MERCADOLIBRE_backup_{timestamp}.xlsx"
        wb_concentrado.save(backup_path)
        print(f"\nBackup creado en: {backup_path}")
        wb_concentrado.save(concentrado_path)
        print("\n=== RESUMEN DE PROCESAMIENTO ===")
        print(f"Total filas actualizadas: {updated_rows}")
        print(f"Total filas nuevas agregadas: {new_rows_count}")
        print(f"Filas duplicadas omitidas: {duplicates_omitted_global}")
        if invalid_files:
            print(f"Archivos no procesados: {len(invalid_files)}")
            for f in invalid_files:
                print(f"- {f}")
        print("================================")
    else:
        print("\nNo se requirieron cambios en el concentrado")
    
    wb_concentrado.close()

if __name__ == "__main__":
    process_excel()#