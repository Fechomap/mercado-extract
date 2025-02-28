import os
import pandas as pd
import pdfplumber
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

class EstadoCuentaProcessor:
    def __init__(self, input_folder="MERCADOPDF", output_folder="MERCADOEXCEL", excel_file="estado_cuenta.xlsx"):
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.excel_file = excel_file
        self.processed_count = 0
        self.error_count = 0
        # Nuevos contadores para el reporte
        self.total_pdfs = 0
        self.pdfs_sin_datos = []
        self.pdfs_con_error = []
        self.pdfs_procesados = []
        self.transacciones_por_pdf = {}
        self.pdfs_con_duplicados = set()
        os.makedirs(output_folder, exist_ok=True)
    
    def get_pdf_paths(self):
        pdf_files = [os.path.join(self.input_folder, f) for f in os.listdir(self.input_folder) if f.endswith(".pdf")]
        if not pdf_files:
            raise FileNotFoundError("No se encontraron archivos PDF en la carpeta MERCADOPDF")
        return pdf_files
    
    def format_date(self, date_str):
        try:
            return datetime.strptime(date_str, '%d-%m-%Y').strftime('%d/%m/%Y')
        except Exception:
            return date_str
    
    def process_line(self, line):
        pattern = r'(\d{2}-\d{2}-\d{4})\s+(.*?)\s+(\d{11})\s+\$\s*([-\d,.]+)\s+\$\s*([-\d,.]+)'
        match = re.search(pattern, line)
        if match:
            fecha, descripcion, id_operacion, valor, saldo = match.groups()
            self.processed_count += 1
            return {
                'Fecha': self.format_date(fecha),
                'Descripción': descripcion.strip(),
                'ID de la operación': id_operacion,
                'Valor': valor.strip(),
                'Saldo': saldo.strip()
            }
        else:
            return None
    
    def process_pdf(self):
        transactions = []
        pdf_paths = self.get_pdf_paths()
        self.total_pdfs = len(pdf_paths)
        print(f"Se encontraron {self.total_pdfs} archivos PDF. Iniciando procesamiento...")
        
        # Procesar cada PDF y asignar un número de orden
        for orden, pdf_path in enumerate(pdf_paths, start=1):
            pdf_name = os.path.basename(pdf_path)
            print(f"\nProcesando archivo: {pdf_path} (Orden: {orden})")
            pdf_transactions = 0
            
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    for page_num, page in enumerate(pdf.pages, 1):
                        print(f"  Procesando página {page_num} de {len(pdf.pages)}")
                        text = page.extract_text()
                        if not text:
                            continue
                        lines = text.split('\n')
                        
                        for line in lines:
                            # Omitir cabeceras y líneas no relevantes
                            if "Fecha" in line and "Descripción" in line:
                                continue
                            if "Fecha de generación:" in line:
                                continue
                            
                            transaction = self.process_line(line)
                            if transaction:
                                # Se agrega el nombre del PDF y el orden de procesamiento
                                transaction['Archivo'] = pdf_name
                                transaction['Orden'] = orden
                                transactions.append(transaction)
                                pdf_transactions += 1
                
                # Registrar información sobre el PDF procesado
                if pdf_transactions > 0:
                    self.pdfs_procesados.append(pdf_name)
                    self.transacciones_por_pdf[pdf_name] = pdf_transactions
                else:
                    self.pdfs_sin_datos.append(pdf_name)
                    
            except Exception as e:
                print(f"Error al abrir el PDF {pdf_path}: {str(e)}")
                self.pdfs_con_error.append(pdf_name)
                self.error_count += 1
        
        if not transactions:
            print("¡Advertencia! No se encontraron transacciones en los PDF.")
            return pd.DataFrame()
        
        # Crear DataFrame con todos los registros
        df = pd.DataFrame(transactions)
        
        # Convertir la columna "Valor" a numérico en una columna auxiliar (para robustez en la comparación)
        df['Valor_num'] = df['Valor'].replace(r'[\$,]', '', regex=True).astype(float)
        
        # Para identificar duplicados entre PDFs
        duplicados_df = df.copy()
        duplicados_por_id = duplicados_df.groupby(['ID de la operación', 'Valor_num']).filter(lambda x: len(x) > 1)
        
        if not duplicados_por_id.empty:
            # Identificar qué PDFs tienen transacciones duplicadas
            archivos_con_duplicados = duplicados_por_id['Archivo'].unique()
            self.pdfs_con_duplicados.update(archivos_con_duplicados)
        
        # Aplicar la lógica de filtrado:
        # Se agrupa por "ID de la operación" y "Valor_num". 
        # Si en un grupo los registros provienen de distintos PDFs (más de un valor único en "Orden"),
        # se conservan únicamente aquellos con el valor mínimo de "Orden".
        # Si todos son del mismo PDF, se mantienen todos.
        df_before = len(df)
        def filtra_grupo(grupo):
            if grupo['Orden'].nunique() == 1:
                return grupo
            else:
                min_orden = grupo['Orden'].min()
                return grupo[grupo['Orden'] == min_orden]
        
        df_filtrado = df.groupby(['ID de la operación', 'Valor_num'], group_keys=False).apply(filtra_grupo)
        descartados = df_before - len(df_filtrado)
        if descartados > 0:
            print(f"Se descartaron {descartados} registros duplicados entre PDFs.")
        
        # Eliminar la columna auxiliar
        df_filtrado = df_filtrado.drop(columns=['Valor_num'])
        
        return df_filtrado
    
    def save_to_excel(self, df):
        if df.empty:
            print("No hay datos para guardar en Excel.")
            return
        
        # Seleccionar únicamente las columnas relevantes y crear una copia para evitar SettingWithCopyWarning
        df_to_save = df[["Fecha", "Descripción", "ID de la operación", "Valor", "Saldo"]].copy()
        excel_path = os.path.join(self.output_folder, self.excel_file)
        
        try:
            date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')
            
            # Conversión de columnas a los tipos deseados usando .loc para evitar SettingWithCopyWarning
            df_to_save.loc[:, 'ID de la operación'] = pd.to_numeric(df_to_save['ID de la operación'], errors='coerce')
            df_to_save.loc[:, 'Valor'] = df_to_save['Valor'].replace(r'[\$,]', '', regex=True).astype(float)
            df_to_save.loc[:, 'Saldo'] = df_to_save['Saldo'].replace(r'[\$,]', '', regex=True).astype(float)
            
            if os.path.exists(excel_path):
                # Caso: Excel existente (se agregan solo los registros nuevos)
                wb = load_workbook(excel_path)
                ws = wb.active
                
                # Extraer las IDs existentes (se asume que la ID de la operación está en la tercera columna)
                existing_ids = set()
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[2] is not None:
                        existing_ids.add(row[2])
                
                df_new = df_to_save[~df_to_save['ID de la operación'].isin(existing_ids)]
                if df_new.empty:
                    print("No hay registros nuevos para agregar. El Excel ya contiene estos datos.")
                    return
                
                start_row = ws.max_row + 1
                for r in dataframe_to_rows(df_new, index=False, header=False):
                    ws.append(r)
                
                # Aplicar formatos a las filas nuevas agregadas
                for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                    cell_fecha = row[0]
                    try:
                        dt = datetime.strptime(cell_fecha.value, '%d/%m/%Y')
                        cell_fecha.value = dt
                        cell_fecha.style = date_style
                    except Exception:
                        pass
                    row[2].number_format = '0'
                    if len(row) > 3:
                        row[3].number_format = '#,##0.00'
                    if len(row) > 4:
                        row[4].number_format = '#,##0.00'
                
                nuevos_registros = len(df_new)
            else:
                # Caso: Nuevo Excel
                wb = Workbook()
                ws = wb.active
                header = list(df_to_save.columns)
                ws.append(header)
                for r in dataframe_to_rows(df_to_save, index=False, header=False):
                    ws.append(r)
                for row in ws.iter_rows(min_row=2):
                    cell_fecha = row[0]
                    try:
                        dt = datetime.strptime(cell_fecha.value, '%d/%m/%Y')
                        cell_fecha.value = dt
                        cell_fecha.style = date_style
                    except Exception:
                        pass
                    row[2].number_format = '0'
                    if len(row) > 3:
                        row[3].number_format = '#,##0.00'
                    if len(row) > 4:
                        row[4].number_format = '#,##0.00'
                nuevos_registros = len(df_to_save)
            
            wb.save(excel_path)
            
            # Generar reporte detallado
            self.generar_reporte_final(nuevos_registros)
            
        except Exception as e:
            print(f"Error al guardar el archivo Excel: {str(e)}")

    def generar_reporte_final(self, nuevos_registros):
        print("\n" + "="*50)
        print("           REPORTE DETALLADO DE PROCESAMIENTO           ")
        print("="*50)
        
        # Resumen general
        print("\n[RESUMEN GENERAL]")
        print(f"- Total de archivos PDF encontrados: {self.total_pdfs}")
        print(f"- PDFs procesados con éxito: {len(self.pdfs_procesados)}")
        print(f"- PDFs sin datos relevantes: {len(self.pdfs_sin_datos)}")
        print(f"- PDFs con errores: {len(self.pdfs_con_error)}")
        print(f"- PDFs con transacciones duplicadas: {len(self.pdfs_con_duplicados)}")
        
        # Detalles de procesamiento
        print("\n[DETALLES DE PROCESAMIENTO]")
        print(f"- Total de registros extraídos: {self.processed_count}")
        print(f"- Registros con error (excepciones): {self.error_count}")
        print(f"- Registros nuevos agregados en este procesamiento: {nuevos_registros}")
        
        # Detalle por PDF
        if self.pdfs_procesados:
            print("\n[PDFS PROCESADOS CON ÉXITO]")
            for pdf in self.pdfs_procesados:
                transacciones = self.transacciones_por_pdf.get(pdf, 0)
                duplicado = " (con duplicados)" if pdf in self.pdfs_con_duplicados else ""
                print(f"- {pdf}: {transacciones} transacciones{duplicado}")
        
        if self.pdfs_sin_datos:
            print("\n[PDFS SIN DATOS RELEVANTES]")
            for pdf in self.pdfs_sin_datos:
                print(f"- {pdf}")
        
        if self.pdfs_con_error:
            print("\n[PDFS CON ERRORES]")
            for pdf in self.pdfs_con_error:
                print(f"- {pdf}")
        
        print("\n[INFORMACIÓN DEL ARCHIVO DE SALIDA]")
        print(f"- Archivo guardado en: {os.path.join(self.output_folder, self.excel_file)}")
        print("="*50)

def main():
    try:
        processor = EstadoCuentaProcessor()
        print("Iniciando procesamiento de los PDF...")
        df = processor.process_pdf()
        print("\nGuardando resultados en Excel...")
        processor.save_to_excel(df)
    except Exception as e:
        print(f"Error en la ejecución del programa: {str(e)}")
        raise

if __name__ == "__main__":
    main()