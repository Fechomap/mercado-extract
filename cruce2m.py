import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import os
from datetime import datetime

def cross_excel_data():
    """
    Cruza datos entre el Excel en MERCADOEXCEL y el Excel CONCENTRADO-MERCADOLIBRE.xlsx
    
    El script busca coincidencias entre los IDs de operación de 11 dígitos que están:
    - En la columna C del Excel en MERCADOEXCEL
    - En la columna I del Excel CONCENTRADO-MERCADOLIBRE.xlsx
    
    Para cada coincidencia, copia la información del Excel de MERCADOEXCEL en las 
    columnas D a H del CONCENTRADO-MERCADOLIBRE.xlsx
    
    Genera un reporte detallado incluyendo los IDs que no fueron encontrados.
    """
    try:
        base_path = Path.cwd()
        print(f"Directorio actual: {base_path}")
        
        # Buscar archivo Excel en la carpeta MERCADOEXCEL
        mercado_excel_dir = base_path.joinpath('MERCADOEXCEL')
        if not mercado_excel_dir.exists():
            print(f"Error: No se encontró la carpeta MERCADOEXCEL en {base_path}")
            return
        
        excel_files = list(mercado_excel_dir.glob('*.xls*'))
        if not excel_files:
            print("No se encontraron archivos Excel en MERCADOEXCEL")
            return
        
        # Tomar el primer archivo Excel encontrado (o se podría especificar el nombre exacto)
        mercado_excel_path = excel_files[0]
        print(f"Usando archivo: {mercado_excel_path.name}")
        
        # Verificar que exista el Excel CONCENTRADO-MERCADOLIBRE.xlsx
        concentrado_path = base_path.joinpath('RESULTADO-FINAL', 'CONCENTRADO-MERCADOLIBRE.xlsx')
        if not concentrado_path.exists():
            print("Error: No se encontró el archivo CONCENTRADO-MERCADOLIBRE.xlsx")
            return
        
        # Cargar el archivo de MERCADOEXCEL
        df_mercado = pd.read_excel(mercado_excel_path)
        
        # Mostrar información detallada sobre las columnas
        print("\n[INFORMACIÓN DEL ARCHIVO ORIGEN]")
        print(f"- Nombre del archivo: {mercado_excel_path.name}")
        print(f"- Total de filas: {len(df_mercado)}")
        print(f"- Total de columnas: {df_mercado.shape[1]}")
        print(f"- Nombres de columnas: {list(df_mercado.columns)}")
        print(f"- Primeros valores en columna C (índice 2): {df_mercado.iloc[:5, 2].tolist()}")
        
        # Verificar que tenga al menos 3 columnas
        if df_mercado.shape[1] < 3:
            print(f"Error: El archivo {mercado_excel_path.name} debe tener al menos 3 columnas")
            return
        
        # Extraer los IDs de operación (columna C - índice 2)
        # Convertir a string y asegurar que sean de 11 dígitos
        df_mercado['ID_LIMPIO'] = df_mercado.iloc[:, 2].astype(str).str.replace(r'\D', '', regex=True)
        ids_validos = df_mercado[df_mercado['ID_LIMPIO'].str.len() == 11]
        
        # Mostrar detalle de limpieza de IDs
        print("\n[DETALLE DE LIMPIEZA DE IDs]")
        print(f"- Total de filas en el archivo: {len(df_mercado)}")
        print(f"- Total de IDs después de limpieza: {len(ids_validos)}")
        print(f"- IDs descartados: {len(df_mercado) - len(ids_validos)}")
        
        if len(ids_validos) < len(df_mercado):
            # Mostrar ejemplos de IDs descartados
            ids_descartados = df_mercado[~df_mercado['ID_LIMPIO'].str.len().eq(11)]
            print("\nEjemplos de IDs descartados (no tienen 11 dígitos):")
            for idx, row in ids_descartados.head(5).iterrows():
                print(f"  Fila {idx+1}: Valor original: '{row.iloc[2]}', Después de limpieza: '{row['ID_LIMPIO']}' (longitud: {len(row['ID_LIMPIO'])})")
        
        if len(ids_validos) == 0:
            print(f"Error: No se encontraron IDs de operación válidos (11 dígitos) en la columna C")
            return
        
        print(f"Se encontraron {len(ids_validos)} IDs de operación válidos en MERCADOEXCEL")
        
        # Cargar el archivo CONCENTRADO-MERCADOLIBRE.xlsx usando openpyxl para modificación
        wb_concentrado = openpyxl.load_workbook(concentrado_path)
        ws_concentrado = wb_concentrado.active
        
        # Preparar datos para transferir (las primeras 5 columnas de MERCADOEXCEL serán mapeadas a columnas D-H)
        data_to_transfer = {}
        for _, row in ids_validos.iterrows():
            id_operacion = row['ID_LIMPIO']
            # Obtener los valores de las primeras 5 columnas (si hay menos, se llenan con None)
            valores = [row.iloc[i] if i < df_mercado.shape[1] else None for i in range(5)]
            data_to_transfer[id_operacion] = valores
        
        # Buscar coincidencias en la columna I del CONCENTRADO-MERCADOLIBRE y actualizar D-H
        coincidencias = 0
        total_filas = 0
        
        for row in range(2, ws_concentrado.max_row + 1):
            total_filas += 1
            # Columna I es la 9
            id_cell = ws_concentrado.cell(row=row, column=9).value
            
            # Limpiar el ID para comparación
            if id_cell is not None:
                id_concentrado = str(id_cell).strip().replace('\xa0', '')
                # Eliminar ".0" al final si existe
                if id_concentrado.endswith('.0'):
                    id_concentrado = id_concentrado[:-2]
                # Eliminar caracteres no numéricos
                id_concentrado = ''.join(filter(str.isdigit, id_concentrado))
                
                # Verificar que tenga 11 dígitos
                if len(id_concentrado) == 11 and id_concentrado in data_to_transfer:
                    coincidencias += 1
                    
                    # Transferir datos a las columnas D-H (columnas 4-8)
                    for i, valor in enumerate(data_to_transfer[id_concentrado]):
                        col_letter = get_column_letter(i + 4)  # D=4, E=5, F=6, G=7, H=8
                        ws_concentrado[f"{col_letter}{row}"] = valor
                        
                        # Aplicar formato numérico si es necesario
                        if isinstance(valor, (int, float)):
                            if isinstance(valor, float) and not valor.is_integer():
                                ws_concentrado[f"{col_letter}{row}"].number_format = "0.00"
                            else:
                                ws_concentrado[f"{col_letter}{row}"].number_format = "0"
        
        # Identificar IDs no encontrados para el reporte
        ids_encontrados = set()
        for row in range(2, ws_concentrado.max_row + 1):
            id_cell = ws_concentrado.cell(row=row, column=9).value
            if id_cell is not None:
                id_concentrado = str(id_cell).strip().replace('\xa0', '')
                if id_concentrado.endswith('.0'):
                    id_concentrado = id_concentrado[:-2]
                id_concentrado = ''.join(filter(str.isdigit, id_concentrado))
                if len(id_concentrado) == 11:
                    ids_encontrados.add(id_concentrado)
        
        # Lista de IDs no encontrados
        ids_no_encontrados = []
        for id_operacion in data_to_transfer.keys():
            if id_operacion not in ids_encontrados:
                ids_no_encontrados.append(id_operacion)
        
        # Guardar listado de IDs no encontrados en un archivo de texto
        reporte_path = base_path / "IDs_no_encontrados.txt"
        with open(reporte_path, 'w') as f:
            f.write(f"=== REPORTE DE IDs NO ENCONTRADOS ===\n")
            f.write(f"Fecha y hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Archivo origen: {mercado_excel_path.name}\n")
            f.write(f"Total IDs no encontrados: {len(ids_no_encontrados)}\n\n")
            if ids_no_encontrados:
                f.write("LISTADO DE IDs NO ENCONTRADOS:\n")
                for i, id_op in enumerate(ids_no_encontrados, 1):
                    # Obtener los datos asociados a este ID
                    valores = data_to_transfer[id_op]
                    valores_str = ', '.join(str(v) if v is not None else 'None' for v in valores)
                    f.write(f"{i}. ID: {id_op} - Datos: {valores_str}\n")
            else:
                f.write("Todos los IDs fueron encontrados y procesados.\n")
        
        # Si se hicieron cambios, guardar el archivo
        if coincidencias > 0:
            # Crear backup
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = concentrado_path.parent / f"CONCENTRADO-MERCADOLIBRE_backup_{timestamp}.xlsx"
            wb_concentrado.save(backup_path)
            print(f"Backup creado en: {backup_path}")
            
            # Guardar cambios
            wb_concentrado.save(concentrado_path)
        
        # Mostrar reporte detallado
        print(f"\n=== REPORTE DETALLADO DE OPERACIÓN ===")
        print(f"Archivo origen: {mercado_excel_path.name}")
        print(f"Archivo destino: {concentrado_path.name}")
        print(f"\n[ESTADÍSTICAS]")
        print(f"- Total de registros en el Excel origen: {len(df_mercado)}")
        print(f"- Total de IDs de operación válidos (11 dígitos): {len(ids_validos)}")
        print(f"- Total de filas en CONCENTRADO-MERCADOLIBRE: {total_filas}")
        print(f"- Coincidencias encontradas y actualizadas: {coincidencias}")
        print(f"- IDs no encontrados: {len(ids_no_encontrados)}")
        
        if ids_no_encontrados:
            print(f"\n[IDs NO ENCONTRADOS]")
            print(f"Se ha generado un archivo detallado en: {reporte_path}")
            print("Listado resumido de IDs no encontrados:")
            for i, id_op in enumerate(ids_no_encontrados[:10], 1):  # Mostrar solo los primeros 10
                print(f"- {id_op}")
            if len(ids_no_encontrados) > 10:
                print(f"... y {len(ids_no_encontrados) - 10} más (ver archivo de reporte)")
        
        print("\n[DETALLES DE OPERACIÓN]")
        if coincidencias > 0:
            print(f"- Se actualizaron {coincidencias} registros")
            print(f"- Se creó un backup en: {backup_path}")
        else:
            print("- No se encontraron coincidencias. No se hicieron cambios.")
        
        print("================================")
        
        wb_concentrado.close()
    
    except Exception as e:
        print(f"Error durante el proceso: {str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    print("Iniciando cruce de datos entre Excel de MERCADOEXCEL y CONCENTRADO-MERCADOLIBRE...")
    cross_excel_data()